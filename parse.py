import os
import sys
import time
import time
# The selenium module
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook, Workbook
from time import sleep

from BeautifulSoup import BeautifulSoup
#======================= Log In credentials
email = "julie@eatinghouse.club"
password = "iloveClover1"
#=======================

driver = webdriver.Chrome("%s/chromedriver" % os.path.dirname(os.path.abspath(__file__)))

row_count = 1
wb_writing = Workbook()
wrt = wb_writing.active

def driver_login(email, password):
    driver.get("https://www.elucify.com/")

    sign_up = driver.find_element_by_xpath("//div[@id='splash-nav-right']/a")
    sign_up.click()

    sleep(3)
    sign_up_gmail = driver.find_element_by_xpath("//button[@id='signup-wall-gmail']")
    sign_up_gmail.click()

    sleep(2)
    email_input = driver.find_element_by_xpath("//input[@name='Email']")
    email_input.send_keys(email)

    next_button = driver.find_element_by_xpath("//input[@id='next']")
    next_button.click()

    sleep(2)
    pass_input = driver.find_element_by_xpath("//input[@id='Passwd']")
    pass_input.send_keys(password)

    submit = driver.find_element_by_xpath("//input[@id = 'signIn']")
    submit.click()
    sleep(3)

def driver_search(name):
    global row_count

    search_field = driver.find_element_by_xpath("//input[@id='company-search-box']")
    search_field.send_keys(name)

    sleep(5)


    try:

        match_form_name = name.strip().lower() + '.'
        print match_form_name
        count = 0
        source = driver.page_source
        soup = BeautifulSoup(source)
        i = 1
        bool = False
        for item in soup('div',{ "class" : 'search-result'}):

            if match_form_name in str(item):
                bool = True
                count = i
            i += 1

        position_str = "[" + str(count) + "]"
        if bool:
            result = driver.find_element_by_xpath("//div[@class='search-result']%s" % position_str)
        else:
            result = driver.find_element_by_xpath("//div[@class='search-result']")

        #result = driver.find_element_by_xpath("//div[@class='search-result'][2]" )
        result.click()
        result.click()
        sleep(3)
        row_count += 1
        wrt["B%d" % row_count] = name

        wb_writing.save("sample.xlsx")
        scrape_page()


    except Exception as e:
        print "No companies matching '%s'" % name
        wrt["B%d" % row_count] = name
        wrt["C%d" % row_count] = "No information available"
        row_count += 1
        wb_writing.save("sample.xlsx")
        print e
        search_field.clear()
        return



    search_field.clear()
    sleep(3)

def scrape_page():
    global row_count

    source = driver.page_source
    soup = BeautifulSoup(source)

    employees = soup.find('div' , {'class' :'company-info-row'})
    employees_text = employees.next.nextSibling.contents[1]
    #print employees_text
    wrt["C%d" % row_count] = "Employees: %s" % employees_text

    logo_wrapper = soup.find('div', {'id' : 'web-tech-logos-wrapper'})
    srcs = []
    if logo_wrapper != None:
        img_div = logo_wrapper.next
        src_text = "https://elucify.com" + img_div.img["src"]

        srcs.append(src_text)

        divs = img_div.findNextSiblings('div')

        for div in divs:
            src_text = "https://elucify.com" + div.img["src"]
            srcs.append(src_text)


    #print srcs

    string_src = ""
    for src in srcs:
        string_src += src + ";  "

    print string_src

    wrt["D%d" % row_count] ="Logos:" + string_src

    wb_writing.save("sample.xlsx")


    #bla = soup.find('div', {'class' : 'company-info-div'})
    #print bla[5]

    for item in soup.findAll('div', {'class':'contact high-confidence'}):
        contact = item.find('div',{'class':'contact-title'}).string
        if ("customer success" in contact.lower() or
                "sales" in contact.lower() or
                "account manager" in contact.lower() or
                "account executive" in contact.lower() or
                "agent" in contact.lower()):

                print contact
                name = item.find('div',{'class':'contact-name'}).string
                link = item.find('div',{'class':'linkedin-url'}).a["href"]
                email = item.find('div',{'class':'contact-email'}).string

                wrt["E%d" % row_count] = contact
                wrt["F%d" % row_count] = name
                wrt["G%d" % row_count] = link
                wrt["H%d" % row_count] = email

                wb_writing.save("sample.xlsx")
                row_count += 1
    for item in soup.findAll('div', {'class':'contact medium-confidence'}):
        contact = item.find('div',{'class':'contact-title'}).string
        if ("customer success" in contact.lower() or
                "sales" in contact.lower() or
                "account manager" in contact.lower() or
                "account executive" in contact.lower() or
                "agent" in contact.lower()):

                print contact
                name = item.find('div',{'class':'contact-name'}).string
                link = item.find('div',{'class':'linkedin-url'}).a["href"]
                email = item.find('div',{'class':'contact-email'}).string

                wrt["E%d" % row_count] = contact
                wrt["F%d" % row_count] = name
                wrt["G%d" % row_count] = link
                wrt["H%d" % row_count] = email

                wb_writing.save("sample.xlsx")
                row_count += 1
    #for item in soup('div',{'class':'company-info-div'}):
    #    if ("customer success" in str(item).lower() or
    #        "sales" in str(item).lower() or
    #        "account manager" in str(item).lower() or
    #        "account executive" in str(item).lower() or
    #        "agent" in str(item).lower()):

    #        bla = soup.find('div', {'class' : 'contact medium-confidence'})
    #        print bla



driver_login(email, password)
#driver_search()


wb = load_workbook('List of companies.xlsx')

main_ws = wb.active

spheres_list = []
for i in range(2,1000):
    if main_ws["A%d" % i].value == None:
        break
    spheres_list.append(main_ws["A%d"%i].value)


for item in spheres_list:
    print item

    wrt["A%d" % row_count] = item
    row_count += 1
    wb_writing.save("sample.xlsx")

    companies_list = []
    try:
        new_ws = wb[item]
        for i in range(1,1000):
            if new_ws["A%d" % i].value == None:
                break
            driver_search(new_ws["A%d" % i].value)
            companies_list.append(new_ws["A%d"%i].value)

    except Exception as e:
        pass


    #print companies_list
